import time
import chromedriver_autoinstaller
from selenium import webdriver
import requests
import openpyxl
from bs4 import BeautifulSoup


BASE_URL = 'https://www.sec.gov/'


def get_cik_code_url():
    section_url = BASE_URL + 'cgi-bin/current?q1=0&q2=0&q3=d'
    response = requests.get(section_url)
    soup = BeautifulSoup(response.content, 'lxml')
    response_bar = soup.find('pre')
    url_elems = response_bar.findAll('a')
    cik_urls = list()
    for index in range(len(url_elems)):
        cik_urls_dict = dict()
        if index % 2 != 0:
            cik_urls_dict['CIK_CODE'] = url_elems[index].text
            cik_urls_dict['URL'] = BASE_URL + url_elems[index].get('href')
            cik_urls.append(cik_urls_dict)
    return cik_urls


def checking_location(cik_urls):
    opt = webdriver.ChromeOptions()
    opt.add_argument("--start-maximized")
    chromedriver_autoinstaller.install()
    driver = webdriver.Chrome(options=opt)
    locations = ['CA', 'NY', 'MA', 'CT', 'TX', 'NJ', 'IL', 'NH']
    id = 1
    for company_cik in cik_urls:
        driver.get(company_cik['URL'])
        time.sleep(2)
        soup = BeautifulSoup(driver.page_source, 'lxml')
        state_location = soup.find('span', {'id': 'stateLocation'}).text
        if state_location in locations:
            print('----------------------------------')
            print('>> Saving to excel ...')
            saving_excel(company_cik['URL'], company_cik['CIK_CODE'], id)
            print(f'{id}  {state_location}', company_cik['URL'])
            id += 1


def create_xlsx():
    wb = openpyxl.Workbook()
    file_name = 'output.xlsx'
    wb.save(file_name)


def saving_excel(url, cik, id):
    file_name = 'output.xlsx'
    exel_file = openpyxl.load_workbook(file_name)
    exel_file.worksheets[0].cell(row=id, column=1).value = cik
    exel_file.worksheets[0].cell(row=id, column=2).value = url
    exel_file.save(file_name)
    exel_file.close()


def main():
    print('>> Creating excel file for output ...')
    create_xlsx()
    print('>> Getting CIK codes urls ...')
    urls = get_cik_code_url()
    print('>> Checking locations for companies ...')
    checking_location(urls)


if __name__ == "__main__":
    main()